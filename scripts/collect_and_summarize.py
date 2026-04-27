import hashlib
import json
import re
import sqlite3
import subprocess
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
import yaml
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config.yaml"
AGG_PROMPT_PATH = ROOT / "prompts" / "aggregation_prompt.md"

SUPPORTED_SUFFIXES = {".txt", ".md", ".html", ".htm", ".json", ".csv"}


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def today_str() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def load_config() -> Dict[str, Any]:
    return yaml.safe_load(CONFIG_PATH.read_text(encoding="utf-8"))


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8", errors="ignore")).hexdigest()


def error_log_path(config: Dict[str, Any]) -> Path:
    reports_dir = ROOT / config.get("output", {}).get("reports_dir", "reports")
    reports_dir.mkdir(parents=True, exist_ok=True)
    return reports_dir / f"error_{today_str()}.log"


def append_error(config: Dict[str, Any], message: str, raw_output: str = "") -> None:
    path = error_log_path(config)
    parts = [f"[{now_str()}] {message}"]
    if raw_output:
        parts.append(raw_output)
    parts.append("-" * 80)
    with path.open("a", encoding="utf-8") as f:
        f.write("\n".join(parts) + "\n")


def init_db(db_path: Path) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(db_path)
    cur = con.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS raw_documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            hash TEXT UNIQUE,
            source_path TEXT,
            ingested_at TEXT,
            content TEXT
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS insights (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_hash TEXT,
            created_at TEXT,
            category TEXT,
            title TEXT,
            summary TEXT,
            facts TEXT,
            url TEXT,
            priority TEXT,
            action TEXT,
            confidence REAL
        )
        """
    )
    con.commit()
    con.close()


def already_ingested(db_path: Path, h: str) -> bool:
    con = sqlite3.connect(db_path)
    cur = con.cursor()
    cur.execute("SELECT 1 FROM raw_documents WHERE hash = ?", (h,))
    row = cur.fetchone()
    con.close()
    return row is not None


def save_raw(db_path: Path, h: str, source_path: str, content: str) -> None:
    con = sqlite3.connect(db_path)
    cur = con.cursor()
    cur.execute(
        "INSERT OR IGNORE INTO raw_documents(hash, source_path, ingested_at, content) VALUES (?, ?, ?, ?)",
        (h, source_path, now_str(), content),
    )
    con.commit()
    con.close()


def save_insights(db_path: Path, doc_hash: str, items: List[Dict[str, Any]]) -> None:
    con = sqlite3.connect(db_path)
    cur = con.cursor()
    for item in items:
        cur.execute(
            """
            INSERT INTO insights(
                doc_hash, created_at, category, title, summary, facts, url, priority, action, confidence
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                doc_hash,
                now_str(),
                str(item.get("category", "")),
                str(item.get("title", "")),
                str(item.get("summary", "")),
                json.dumps(item.get("facts", []), ensure_ascii=False),
                str(item.get("url", "")),
                str(item.get("priority", "")),
                str(item.get("action", "")),
                float(item.get("confidence", 0.0) or 0.0),
            ),
        )
    con.commit()
    con.close()


def read_file_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md", ".json", ".csv"}:
        return path.read_text(encoding="utf-8", errors="ignore")
    if suffix in {".html", ".htm"}:
        soup = BeautifulSoup(path.read_text(encoding="utf-8", errors="ignore"), "html.parser")
        return soup.get_text("\n", strip=True)
    return ""


def collect_manual_inputs(config: Dict[str, Any]) -> List[Dict[str, str]]:
    docs: List[Dict[str, str]] = []
    folders = config.get("manual_inputs", {}).get("folders", [])
    for folder in folders:
        folder_path = ROOT / folder
        if not folder_path.exists():
            continue
        for path in folder_path.rglob("*"):
            if path.is_file() and path.suffix.lower() in SUPPORTED_SUFFIXES:
                content = read_file_text(path)
                if content.strip():
                    docs.append({"source_path": str(path), "content": content})
    return docs


def chunk_text(text: str, max_chars: int) -> List[str]:
    text = text.strip()
    if not text:
        return []
    if len(text) <= max_chars:
        return [text]
    return [text[i : i + max_chars] for i in range(0, len(text), max_chars)]


def extract_json_array(text: str) -> List[Dict[str, Any]]:
    try:
        data = json.loads(text)
        if isinstance(data, list):
            return data
        if isinstance(data, dict) and isinstance(data.get("items"), list):
            return data["items"]
    except json.JSONDecodeError:
        pass

    code_block = re.search(r"```(?:json)?\s*(\[.*?\])\s*```", text, flags=re.DOTALL | re.IGNORECASE)
    if code_block:
        return json.loads(code_block.group(1))

    match = re.search(r"\[\s*\{.*\}\s*\]", text, re.DOTALL)
    if match:
        return json.loads(match.group(0))

    raise ValueError("JSON array not found")


def run_gemini(prompt: str, context: str, config: Dict[str, Any]) -> str:
    command = config.get("gemini", {}).get("command", "gemini")
    timeout = int(config.get("gemini", {}).get("timeout_seconds", 180))
    result = subprocess.run(
        [command, "-p", prompt],
        input=context,
        text=True,
        capture_output=True,
        encoding="utf-8",
        errors="ignore",
        timeout=timeout,
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or "gemini command failed")
    return result.stdout.strip()


def load_text_prompt(path: Path, fallback: str) -> str:
    if path.exists():
        return path.read_text(encoding="utf-8")
    return fallback


def insight_prompt() -> str:
    fallback = (
        "入力テキストから重要情報を抽出し、JSON配列のみ返してください。"
        "priorityはA/B/C、Aは即金・案件・開発・自動化に直結です。"
    )
    return load_text_prompt(AGG_PROMPT_PATH, fallback)


def daily_research_prompt(topic_name: str, query: str) -> str:
    base = ROOT / "prompts" / "daily_research_prompt.md"
    template = load_text_prompt(base, "JSON配列のみ返してください。")
    return f"{template}\n\nテーマ: {topic_name}\n調査テーマ:\n{query}\n"


def process_manual_docs(config: Dict[str, Any], db_path: Path) -> None:
    max_chars = int(config.get("gemini", {}).get("max_chars_per_batch", 50000))
    for doc in collect_manual_inputs(config):
        source_path = doc["source_path"]
        content = doc["content"]
        h = sha256_text(source_path + content)
        if already_ingested(db_path, h):
            continue

        print(f"[ingest] {source_path}")
        save_raw(db_path, h, source_path, content)

        for i, chunk in enumerate(chunk_text(content, max_chars)):
            try:
                output = run_gemini(insight_prompt(), chunk, config)
                insights = extract_json_array(output)
                save_insights(db_path, h, insights)
            except Exception as exc:
                append_error(
                    config,
                    f"insight extraction failed: source={source_path} chunk={i} error={exc}",
                    raw_output=output if "output" in locals() else "",
                )
                print(f"[ERROR] insight extraction failed: {source_path} chunk={i}")


def run_daily_research(config: Dict[str, Any], db_path: Path) -> None:
    if not config.get("daily_research", {}).get("enabled", True):
        return

    for topic in config.get("daily_research", {}).get("topics", []):
        name = str(topic.get("name", ""))
        query = str(topic.get("query", ""))
        if not query.strip():
            continue

        print(f"[research] {name}")
        prompt = daily_research_prompt(name, query)
        context = f"実行日: {today_str()}\nテーマ: {name}\n"
        output = ""
        try:
            output = run_gemini(prompt, context, config)
            insights = extract_json_array(output)
            h = sha256_text(name + query + today_str())
            save_raw(db_path, h, f"daily_research:{name}", output)
            save_insights(db_path, h, insights)
        except Exception as exc:
            append_error(config, f"daily research failed: topic={name} error={exc}", raw_output=output)
            print(f"[ERROR] daily research failed: {name}")


def load_today_insights(db_path: Path) -> pd.DataFrame:
    con = sqlite3.connect(db_path)
    today = today_str()
    df = pd.read_sql_query(
        """
        SELECT id, created_at, priority, category, title, summary, facts, action, url, confidence
        FROM insights
        WHERE created_at LIKE ?
        ORDER BY CASE priority WHEN 'A' THEN 1 WHEN 'B' THEN 2 ELSE 3 END, category, id
        """,
        con,
        params=(f"{today}%",),
    )
    con.close()
    return df


def dedupe_for_report(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    df = df.copy()
    df["_url_norm"] = df["url"].fillna("").str.strip().str.lower()
    df["_title_norm"] = df["title"].fillna("").str.strip().str.lower()

    deduped_rows = []
    seen = set()
    for _, row in df.iterrows():
        key = row["_url_norm"] if row["_url_norm"] else row["_title_norm"]
        if not key:
            key = f"id:{row['id']}"
        if key in seen:
            continue
        seen.add(key)
        deduped_rows.append(row)

    out = pd.DataFrame(deduped_rows).drop(columns=["_url_norm", "_title_norm"], errors="ignore")
    return out.reset_index(drop=True)


def write_markdown(df: pd.DataFrame, md_path: Path, date: str) -> None:
    if df.empty:
        md_path.write_text(f"# Daily Summary {date}\n\n本日の新規抽出情報はありません。\n", encoding="utf-8")
        return

    lines = [f"# Daily Summary {date}", ""]
    for priority in ["A", "B", "C"]:
        subset = df[df["priority"] == priority]
        lines.extend([f"## 優先度 {priority}", ""])
        if subset.empty:
            lines.extend(["- 該当なし", ""])
            continue
        for _, row in subset.iterrows():
            lines.append(f"### {row['title']}")
            lines.append(f"- カテゴリ: {row['category']}")
            lines.append(f"- 要約: {row['summary']}")
            if str(row.get("action", "")).strip():
                lines.append(f"- 次アクション: {row['action']}")
            if str(row.get("url", "")).strip():
                lines.append(f"- URL: {row['url']}")
            lines.append("")

    md_path.write_text("\n".join(lines), encoding="utf-8")


def write_excel(df: pd.DataFrame, xlsx_path: Path) -> None:
    export_cols = ["created_at", "priority", "category", "title", "summary", "facts", "action", "url", "confidence"]
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df[export_cols].to_excel(writer, index=False, sheet_name="Daily_Summary")
        ws = writer.book["Daily_Summary"]

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            priority_cell = row[1]
            url_cell = row[7]
            if str(priority_cell.value).strip() == "A":
                for c in row:
                    c.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    c.font = Font(bold=True)

            if url_cell.value and str(url_cell.value).startswith(("http://", "https://")):
                url_cell.hyperlink = str(url_cell.value)
                url_cell.style = "Hyperlink"

        for column_cells in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(12, max_len + 2), 60)


def create_reports(config: Dict[str, Any], db_path: Path) -> None:
    reports_dir = ROOT / config.get("output", {}).get("reports_dir", "reports")
    reports_dir.mkdir(parents=True, exist_ok=True)

    date = today_str()
    md_path = reports_dir / f"daily_summary_{date}.md"
    xlsx_path = reports_dir / f"daily_summary_{date}.xlsx"
    action_path = reports_dir / "important_actions.md"

    df = dedupe_for_report(load_today_insights(db_path))
    write_markdown(df, md_path, date)
    write_excel(df, xlsx_path)

    action_lines = [f"# Important Actions - updated {now_str()}", ""]
    a_df = df[df["priority"] == "A"] if not df.empty else pd.DataFrame()
    if a_df.empty:
        action_lines.append("- 本日のA優先度アクションはありません。")
    else:
        for _, row in a_df.iterrows():
            action_lines.append(f"- [{row['category']}] {row['title']} / {row['action']} / {row['url']}")
    action_path.write_text("\n".join(action_lines), encoding="utf-8")

    print(f"[report] {md_path}")
    print(f"[report] {xlsx_path}")


def sync_to_drive(config: Dict[str, Any]) -> None:
    sync = config.get("sync", {})
    if not sync.get("enabled", False):
        return

    dst = sync.get("google_drive_reports_dir")
    if not dst:
        return

    src = ROOT / config.get("output", {}).get("reports_dir", "reports")
    subprocess.run(["robocopy", str(src), str(dst), "/MIR"], check=False, shell=True)


def main() -> None:
    config = load_config()
    db_path = ROOT / config.get("output", {}).get("database_path", "data/aggregator.sqlite")

    init_db(db_path)
    process_manual_docs(config, db_path)
    run_daily_research(config, db_path)
    create_reports(config, db_path)
    sync_to_drive(config)
    print("[done]")


if __name__ == "__main__":
    main()
